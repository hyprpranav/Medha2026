"""
MEDHA Command Center — Data Upload Script
Reads team data from Excel files and uploads to Firebase Firestore.

Prerequisites:
    pip install openpyxl firebase-admin

Usage:
    1. Place your Firebase service account key JSON file in this directory
    2. Update FIREBASE_CREDENTIALS_FILE and FIREBASE_PROJECT_ID below
    3. Run: python upload_teams.py

This script:
    - Reads both Day 1 and Day 2 Excel registration files
    - Deduplicates teams by teamName (lowercased)
    - Parses members, leader info, project details
    - Uploads to Firestore 'teams' collection
    - Creates initial 'settings' document
"""

import os
import re
import sys
from datetime import datetime

import openpyxl

# ───────────────────────────────────────────
# CONFIGURATION — UPDATE THESE VALUES
# ───────────────────────────────────────────

# Path to your Firebase service account key JSON
FIREBASE_CREDENTIALS_FILE = "serviceAccountKey.json"

# Your Firebase project ID
FIREBASE_PROJECT_ID = "medha-core"

# Path to the details_for_ai folder with Excel files
EXCEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "details_for_ai")

# ───────────────────────────────────────────

def parse_phone(val):
    """Clean phone number to string."""
    if val is None:
        return ""
    s = str(val).strip().replace(" ", "").replace("+91", "")
    # Remove .0 from float conversion
    if s.endswith(".0"):
        s = s[:-2]
    return s


def parse_members(row, headers):
    """Extract members list from a row."""
    members = []

    # Member 1
    name1_idx = next((i for i, h in enumerate(headers) if h and "Name of the Member 1" in str(h)), None)
    dept1_idx = next((i for i, h in enumerate(headers) if h and "Year & Dept" in str(h) and "2" not in str(h) and "3" not in str(h) and "4" not in str(h)), None)

    # Member 2
    name2_idx = next((i for i, h in enumerate(headers) if h and "Name of the Member 2" in str(h)), None)
    dept2_idx = next((i for i, h in enumerate(headers) if h and "Eg: IV - ECE 2" in str(h)), None)

    # Member 3
    name3_idx = next((i for i, h in enumerate(headers) if h and "Name of the Member 3" in str(h)), None)
    dept3_idx = next((i for i, h in enumerate(headers) if h and "Eg: IV - ECE 3" in str(h)), None)

    # Member 4
    name4_idx = next((i for i, h in enumerate(headers) if h and "Name of the Member 4" in str(h)), None)
    dept4_idx = next((i for i, h in enumerate(headers) if h and "Eg: IV - ECE 4" in str(h)), None)

    # Extract members by index
    member_pairs = [
        (name1_idx, dept1_idx),
        (name2_idx, dept2_idx),
        (name3_idx, dept3_idx),
        (name4_idx, dept4_idx),
    ]

    for name_idx, dept_idx in member_pairs:
        if name_idx is not None and name_idx < len(row):
            name = row[name_idx]
            if name and str(name).strip() and str(name).strip() not in ["-", "No", "None", "Nil"]:
                dept = ""
                if dept_idx is not None and dept_idx < len(row):
                    dept = str(row[dept_idx] or "").strip()
                    if dept in ["-", "No", "None", "Nil"]:
                        dept = ""
                members.append({
                    "name": str(name).strip(),
                    "dept": dept,
                })

    return members


def read_excel_files():
    """Read all Excel files and return parsed team data."""
    teams = {}  # keyed by teamName_lower to deduplicate

    for fname in os.listdir(EXCEL_DIR):
        if not fname.endswith(".xlsx"):
            continue

        fpath = os.path.join(EXCEL_DIR, fname)
        print(f"\n📁 Reading: {fname}")

        wb = openpyxl.load_workbook(fpath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        # Find column indices
        def find_col(keyword):
            for i, h in enumerate(headers):
                if h and keyword.lower() in str(h).lower():
                    return i
            return None

        col_team = find_col("Team Name")
        col_leader = find_col("Team Leader Name")
        col_phone = find_col("Phone Number")
        col_alt_phone = find_col("Alternative Phone")
        col_email_id = find_col("Email id")
        col_college = find_col("College Name")
        col_district = find_col("District")
        col_state = find_col("State")
        col_boys = find_col("Team Boys Count")
        col_girls = find_col("Team Girls Count")
        col_track = find_col("Problem Statement Track")
        col_track1 = find_col("Track 1")
        col_track2 = find_col("Track 2")
        col_project = find_col("Title of the project")
        col_accommodation = find_col("Accommodation")
        col_transaction = find_col("Transaction id")

        for row_data in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            row = list(row_data)

            team_name = str(row[col_team] or "").strip() if col_team is not None else ""
            if not team_name:
                continue

            team_key = team_name.lower().strip()

            # Skip if already added (deduplicate)
            if team_key in teams:
                print(f"  ⚠ Duplicate skipped: {team_name}")
                continue

            leader_name = str(row[col_leader] or "").strip() if col_leader is not None else ""
            leader_phone = parse_phone(row[col_phone]) if col_phone is not None else ""
            leader_email = str(row[col_email_id] or "").strip() if col_email_id is not None else ""
            college = str(row[col_college] or "").strip() if col_college is not None else ""
            district = str(row[col_district] or "").strip() if col_district is not None else ""
            state = str(row[col_state] or "").strip() if col_state is not None else ""

            boys = int(row[col_boys] or 0) if col_boys is not None and row[col_boys] else 0
            girls = int(row[col_girls] or 0) if col_girls is not None and row[col_girls] else 0

            # Track
            track = str(row[col_track] or "").strip() if col_track is not None else ""
            sub_track = ""
            if col_track1 is not None and row[col_track1]:
                sub_track = str(row[col_track1]).strip()
            elif col_track2 is not None and row[col_track2]:
                sub_track = str(row[col_track2]).strip()

            project = str(row[col_project] or "").strip() if col_project is not None else ""
            accommodation = str(row[col_accommodation] or "").strip() if col_accommodation is not None else ""
            transaction_id = str(row[col_transaction] or "").strip() if col_transaction is not None else ""

            members = parse_members(row, headers)

            total_members = len(members)
            if total_members == 0:
                total_members = boys + girls

            teams[team_key] = {
                "teamName": team_name,
                "teamNameLower": team_key,
                "leaderName": leader_name,
                "leaderEmail": leader_email.rstrip("@"),  # Fix trailing @
                "leaderPhone": leader_phone,
                "collegeName": college,
                "district": district,
                "state": state,
                "members": members,
                "totalMembers": total_members,
                "boysCount": boys,
                "girlsCount": girls,
                "track": f"{track} — {sub_track}" if sub_track else track,
                "projectTitle": project,
                "accommodation": accommodation.lower() == "yes",
                "transactionId": transaction_id,
                # Attendance fields — initialized empty
                "attendanceStatus": None,
                "presentCount": None,
                "absentCount": None,
                "checkedIn": False,
                "checkedInBy": None,
                "checkedInByName": None,
                "checkedInAt": None,
                "attendanceRound": None,
                "attendanceLocked": False,
                "attendanceRecords": [],
                # QR fields
                "qrToken": None,
                "qrGeneratedAt": None,
                # Timestamps
                "createdAt": datetime.now().isoformat(),
                "lastModified": datetime.now().isoformat(),
                "source": fname,
            }

        wb.close()

    return teams


def upload_to_firestore(teams):
    """Upload parsed teams to Firestore."""
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore
    except ImportError:
        print("\n❌ firebase-admin not installed!")
        print("   Run: pip install firebase-admin")
        sys.exit(1)

    cred_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), FIREBASE_CREDENTIALS_FILE)

    if not os.path.exists(cred_path):
        print(f"\n❌ Service account key not found: {cred_path}")
        print("   Download it from Firebase Console → Project Settings → Service Accounts")
        print("   Save it as 'serviceAccountKey.json' in the scripts/ folder")
        sys.exit(1)

    # Initialize Firebase Admin
    cred = credentials.Certificate(cred_path)
    firebase_admin.initialize_app(cred, {"projectId": FIREBASE_PROJECT_ID})
    db = firestore.client()

    print(f"\n🔥 Uploading {len(teams)} teams to Firestore...\n")

    batch = db.batch()
    count = 0

    for team_key, team_data in teams.items():
        # Use auto-generated IDs for teams
        doc_ref = db.collection("teams").document()
        batch.set(doc_ref, team_data)
        count += 1
        print(f"  ✅ {count}. {team_data['teamName']} ({team_data['collegeName'][:40]})")

        # Commit in batches of 400 (Firestore limit is 500)
        if count % 400 == 0:
            batch.commit()
            batch = db.batch()

    # Commit remaining
    if count % 400 != 0:
        batch.commit()

    print(f"\n✅ Successfully uploaded {count} teams to Firestore!")

    # Create settings document
    settings_ref = db.collection("settings").document("main")
    settings_ref.set({
        "attendanceEnabled": False,
        "currentSession": "Morning",
        "sessions": ["Morning", "Afternoon", "Final"],
        "createdAt": firestore.SERVER_TIMESTAMP,
    }, merge=True)
    print("✅ Settings document created/updated")

    # Create admin user document
    admin_email = "harishpranavs259@gmail.com"
    print(f"ℹ  Master admin email: {admin_email}")
    print("   Admin role will be auto-assigned on first login with this email.")


def export_preview(teams):
    """Print preview of parsed teams without uploading."""
    print(f"\n📊 Parsed {len(teams)} unique teams:\n")
    for i, (key, team) in enumerate(teams.items(), 1):
        print(f"  {i:>2}. {team['teamName']:<30s} | {team['collegeName'][:35]:<35s} | Members: {team['totalMembers']} | {team['track'][:25]}")
    print()


if __name__ == "__main__":
    teams = read_excel_files()
    export_preview(teams)

    if "--upload" in sys.argv:
        upload_to_firestore(teams)
    else:
        print("=" * 60)
        print("DRY RUN — No data was uploaded.")
        print("To upload to Firestore, run:")
        print("  python upload_teams.py --upload")
        print()
        print("Before uploading, make sure:")
        print("  1. serviceAccountKey.json is in this folder")
        print("  2. FIREBASE_PROJECT_ID is set correctly in the script")
        print("=" * 60)
