import openpyxl
import os

folder = os.path.dirname(os.path.abspath(__file__))
details_dir = os.path.join(folder, "details_for_ai")

for fname in os.listdir(details_dir):
    if fname.endswith('.xlsx'):
        fpath = os.path.join(details_dir, fname)
        print(f"\n{'='*80}")
        print(f"FILE: {fname}")
        print(f"{'='*80}")
        wb = openpyxl.load_workbook(fpath)
        ws = wb.active
        print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")
        
        # Print headers
        headers = [cell.value for cell in ws[1]]
        print(f"\nHEADERS: {headers}")
        
        # Print all rows
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
            print(f"\nRow {i}: {list(row)}")
        wb.close()
